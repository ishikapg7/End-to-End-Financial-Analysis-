"""
Financial Analysis Chatbot — Streamlit Web App
Run with:  streamlit run app.py
"""

import os
import io
import json
import re
import tempfile
from datetime import datetime
from pathlib import Path

import streamlit as st
import anthropic

# ── optional PDF support ──────────────────────────────────────────────────────
try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# ── Excel support ─────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ─────────────────────────────────────────────────────────────────────────────
MODEL     = "claude-opus-4-6"
DARK_BLUE  = "1F3864";  MID_BLUE  = "2F5496"; LIGHT_BLUE = "BDD7EE"
WHITE      = "FFFFFF";  LIGHT_GREY = "F2F2F2"; NEG_RED   = "FF0000"

# ═════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ═════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="10-K Financial Analyst",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  /* Chat bubbles */
  [data-testid="stChatMessage"] { border-radius: 12px; margin-bottom: 8px; }

  /* Sidebar */
  section[data-testid="stSidebar"] { background: #0f1b2d; }
  section[data-testid="stSidebar"] * { color: #e0e8f0 !important; }
  section[data-testid="stSidebar"] .stButton button {
      background: #2f5496; color: white; border-radius: 8px;
      width: 100%; border: none; font-weight: 600;
  }
  section[data-testid="stSidebar"] .stButton button:hover { background: #1f3864; }

  /* Remove top padding */
  .block-container { padding-top: 1.5rem; }

  /* Status badge */
  .badge { display:inline-block; padding:2px 10px; border-radius:12px;
           font-size:12px; font-weight:600; margin:2px; }
  .badge-green  { background:#d4edda; color:#155724; }
  .badge-orange { background:#fff3cd; color:#856404; }

  /* Analysis card */
  .analysis-card {
      background: #f8faff; border-left: 4px solid #2f5496;
      border-radius: 8px; padding: 16px; margin: 8px 0;
  }
  .metric-row { display:flex; gap:24px; flex-wrap:wrap; margin-top:8px; }
  .metric { text-align:center; }
  .metric-val { font-size:22px; font-weight:700; color:#1f3864; }
  .metric-lbl { font-size:11px; color:#666; }
</style>
""", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ═════════════════════════════════════════════════════════════════════════════
def _init_state():
    defaults = dict(
        messages=[],           # chat history [{role, content}]
        companies=[],          # [{name, text}]
        extracted=[],          # raw JSON from Claude
        analyses=[],           # KPI dicts
        dcf_results=[],        # DCF dicts
        excel_bytes=None,      # bytes for download
        analysis_done=False,
        api_key="",
    )
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_state()


def get_client():
    key = st.session_state.api_key or os.getenv("ANTHROPIC_API_KEY", "")
    if not key:
        return None
    return anthropic.Anthropic(api_key=key)


# ═════════════════════════════════════════════════════════════════════════════
# FINANCIAL LOGIC  (same as CLI version)
# ═════════════════════════════════════════════════════════════════════════════

def read_uploaded_file(uploaded) -> str:
    if uploaded.name.lower().endswith(".pdf"):
        if not PDF_SUPPORT:
            st.error("pdfplumber not installed. Run: pip install pdfplumber")
            return ""
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded.read())
            tmp_path = tmp.name
        text_parts = []
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        os.unlink(tmp_path)
        return "\n".join(text_parts)
    return uploaded.read().decode("utf-8", errors="replace")


EXTRACTION_SYSTEM = """
You are a senior financial analyst specialising in public-company 10-K filings.
Extract ALL numerical financial data with extreme precision.
Return ONLY valid JSON — no markdown fences, no commentary.
Use millions (USD) for monetary figures. Use % for margins/rates.
If a value is not found write null.
"""

EXTRACTION_PROMPT = """
Extract the following from this 10-K filing and return JSON matching EXACTLY
this schema (values in USD millions unless noted):

{
  "company_name": "string",
  "ticker": "string",
  "fiscal_year": "YYYY",
  "currency": "USD",

  "income_statement": {
    "revenue": [Y-2, Y-1, Y0],
    "cost_of_goods_sold": [Y-2, Y-1, Y0],
    "gross_profit": [Y-2, Y-1, Y0],
    "rd_expense": [Y-2, Y-1, Y0],
    "sg_a_expense": [Y-2, Y-1, Y0],
    "operating_income": [Y-2, Y-1, Y0],
    "ebitda": [Y-2, Y-1, Y0],
    "interest_expense": [Y-2, Y-1, Y0],
    "pretax_income": [Y-2, Y-1, Y0],
    "income_tax": [Y-2, Y-1, Y0],
    "net_income": [Y-2, Y-1, Y0],
    "eps_diluted": [Y-2, Y-1, Y0],
    "shares_diluted": [Y-2, Y-1, Y0]
  },

  "balance_sheet": {
    "cash_and_equivalents": [Y-1, Y0],
    "accounts_receivable": [Y-1, Y0],
    "inventory": [Y-1, Y0],
    "total_current_assets": [Y-1, Y0],
    "ppe_net": [Y-1, Y0],
    "intangible_assets": [Y-1, Y0],
    "goodwill": [Y-1, Y0],
    "total_assets": [Y-1, Y0],
    "accounts_payable": [Y-1, Y0],
    "short_term_debt": [Y-1, Y0],
    "total_current_liabilities": [Y-1, Y0],
    "long_term_debt": [Y-1, Y0],
    "total_liabilities": [Y-1, Y0],
    "total_equity": [Y-1, Y0],
    "retained_earnings": [Y-1, Y0]
  },

  "cash_flow": {
    "cfo": [Y-2, Y-1, Y0],
    "capex": [Y-2, Y-1, Y0],
    "free_cash_flow": [Y-2, Y-1, Y0],
    "cfi": [Y-2, Y-1, Y0],
    "cff": [Y-2, Y-1, Y0],
    "net_change_in_cash": [Y-2, Y-1, Y0],
    "depreciation_amortization": [Y-2, Y-1, Y0],
    "stock_based_compensation": [Y-2, Y-1, Y0]
  },

  "dcf_inputs": {
    "beta": null,
    "debt_outstanding": null,
    "shares_outstanding": null,
    "effective_tax_rate": null,
    "cost_of_debt": null,
    "risk_free_rate": 4.5,
    "equity_risk_premium": 5.5,
    "terminal_growth_rate": 2.5,
    "revenue_growth_estimate_y1": null,
    "revenue_growth_estimate_y2": null,
    "revenue_growth_estimate_y3": null,
    "revenue_growth_estimate_y4": null,
    "revenue_growth_estimate_y5": null,
    "ebitda_margin_estimate": null,
    "da_pct_revenue": null,
    "capex_pct_revenue": null,
    "nwc_change_pct_revenue": null
  }
}

10-K TEXT (first 120,000 chars):
"""


def safe(lst, idx, default=None):
    try:
        v = lst[idx]
        return v if v is not None else default
    except (IndexError, TypeError):
        return default


def pct(num, den, decimals=1):
    try:
        return round((num / den) * 100, decimals)
    except (ZeroDivisionError, TypeError):
        return None


def _parse_json_from_text(raw: str) -> dict:
    """Try multiple strategies to extract a JSON object from Claude's response."""
    # 1. Strip markdown fences
    cleaned = re.sub(r"```(?:json)?", "", raw).strip().rstrip("`").strip()

    # 2. Direct parse
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # 3. Find the largest {...} block
    candidates = []
    for m in re.finditer(r"\{", cleaned):
        start = m.start()
        depth = 0
        for i, ch in enumerate(cleaned[start:], start):
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    candidates.append(cleaned[start : i + 1])
                    break

    # try largest first
    for candidate in sorted(candidates, key=len, reverse=True):
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            continue

    raise ValueError("No valid JSON object found in response.")


def extract_financial_data(name: str, text: str, client) -> dict:
    # Try with a large chunk first; fall back to smaller if needed
    for chunk_size, max_tok in [(120_000, 8000), (60_000, 6000)]:
        truncated = text[:chunk_size]
        try:
            with client.messages.stream(
                model=MODEL,
                max_tokens=max_tok,
                thinking={"type": "adaptive"},
                system=EXTRACTION_SYSTEM,
                messages=[{
                    "role": "user",
                    "content": (
                        EXTRACTION_PROMPT + truncated
                        + "\n\nIMPORTANT: Return ONLY the raw JSON object. "
                        "No explanation, no markdown, no code fences. "
                        "Start your response with { and end with }."
                    ),
                }],
            ) as stream:
                response = stream.get_final_message()

            # Collect all text blocks (thinking blocks are skipped automatically)
            raw = " ".join(
                b.text for b in response.content if b.type == "text"
            )

            data = _parse_json_from_text(raw)
            data.setdefault("company_name", name)
            return data

        except (ValueError, json.JSONDecodeError) as e:
            if chunk_size == 60_000:
                raise ValueError(
                    f"Could not extract valid JSON for '{name}' after 2 attempts. "
                    f"Last error: {e}"
                )
            # retry with smaller chunk
            continue


def three_statement_analysis(data: dict) -> dict:
    inc = data.get("income_statement", {})
    bs  = data.get("balance_sheet", {})
    cf  = data.get("cash_flow", {})
    rev0 = safe(inc.get("revenue"), 2)
    rev1 = safe(inc.get("revenue"), 1)
    return {
        "company":            data.get("company_name"),
        "ticker":             data.get("ticker"),
        "year":               data.get("fiscal_year"),
        "revenue_y0":         rev0,
        "revenue_yoy_pct":    pct(rev0 - rev1, rev1) if rev0 and rev1 else None,
        "gross_margin_pct":   pct(safe(inc.get("gross_profit"), 2), rev0),
        "ebitda_margin_pct":  pct(safe(inc.get("ebitda"), 2), rev0),
        "operating_margin_pct": pct(safe(inc.get("operating_income"), 2), rev0),
        "net_margin_pct":     pct(safe(inc.get("net_income"), 2), rev0),
        "eps_diluted":        safe(inc.get("eps_diluted"), 2),
        "current_ratio":      (
            (safe(bs.get("total_current_assets"), 1) or 0) /
            (safe(bs.get("total_current_liabilities"), 1) or 1)
            if safe(bs.get("total_current_assets"), 1) else None
        ),
        "debt_to_equity":     (
            (safe(bs.get("long_term_debt"), 1) or 0) /
            (safe(bs.get("total_equity"), 1) or 1)
            if safe(bs.get("total_equity"), 1) else None
        ),
        "net_debt":           (
            (safe(bs.get("long_term_debt"), 1) or 0)
            + (safe(bs.get("short_term_debt"), 1) or 0)
            - (safe(bs.get("cash_and_equivalents"), 1) or 0)
        ),
        "roe_pct":  pct(safe(inc.get("net_income"), 2), safe(bs.get("total_equity"), 1)),
        "roa_pct":  pct(safe(inc.get("net_income"), 2), safe(bs.get("total_assets"), 1)),
        "fcf_y0":   safe(cf.get("free_cash_flow"), 2),
        "fcf_margin": pct(safe(cf.get("free_cash_flow"), 2), rev0),
        "capex_y0": safe(cf.get("capex"), 2),
        "cfo_y0":   safe(cf.get("cfo"), 2),
    }


def build_dcf(data: dict) -> dict:
    inc    = data.get("income_statement", {})
    inputs = data.get("dcf_inputs", {})
    base_rev   = safe(inc.get("revenue"), 2) or 0
    base_ebitda = safe(inc.get("ebitda"), 2) or 0
    base_em     = (base_ebitda / base_rev) if base_rev else 0.20

    beta   = inputs.get("beta") or 1.1
    rfr    = (inputs.get("risk_free_rate") or 4.5) / 100
    erp    = (inputs.get("equity_risk_premium") or 5.5) / 100
    cost_e = rfr + beta * erp

    debt   = inputs.get("debt_outstanding") or safe(data.get("balance_sheet", {}).get("long_term_debt"), 1) or 0
    equity = safe(data.get("balance_sheet", {}).get("total_equity"), 1) or base_rev
    total  = debt + equity
    tax_r  = (inputs.get("effective_tax_rate") or 21) / 100
    cost_d = (inputs.get("cost_of_debt") or 5.0) / 100
    wd = debt / total if total else 0.30
    we = equity / total if total else 0.70
    wacc = we * cost_e + wd * cost_d * (1 - tax_r)

    g = [(inputs.get(f"revenue_growth_estimate_y{i}") or None) for i in range(1, 6)]
    defaults = [8.0, 7.0, 6.0, 5.0, 4.0]
    g = [(v if v is not None else defaults[i]) / 100 for i, v in enumerate(g)]

    ebitda_m  = (inputs.get("ebitda_margin_estimate") or base_em * 100) / 100
    da_pct    = (inputs.get("da_pct_revenue") or 3.0) / 100
    capex_pct = (inputs.get("capex_pct_revenue") or 4.0) / 100
    nwc_pct   = (inputs.get("nwc_change_pct_revenue") or 1.5) / 100
    tgr       = (inputs.get("terminal_growth_rate") or 2.5) / 100

    years, revenues, ebitdas, fcfs, pv_fcfs = [], [], [], [], []
    rev = base_rev
    for i, gr in enumerate(g, 1):
        rev       = rev * (1 + gr)
        ebitda_v  = rev * ebitda_m
        da_v      = rev * da_pct
        ebit_v    = ebitda_v - da_v
        nopat     = ebit_v * (1 - tax_r)
        capex_v   = rev * capex_pct
        nwc_ch    = rev * nwc_pct
        fcf_v     = nopat + da_v - capex_v - nwc_ch
        pv        = fcf_v / ((1 + wacc) ** i)
        years.append(i); revenues.append(round(rev, 1))
        ebitdas.append(round(ebitda_v, 1)); fcfs.append(round(fcf_v, 1))
        pv_fcfs.append(round(pv, 1))

    tv_fcf  = fcfs[-1] * (1 + tgr) / (wacc - tgr) if wacc > tgr else 0
    pv_tv   = tv_fcf / ((1 + wacc) ** 5)
    sum_pv  = sum(pv_fcfs)
    ev      = sum_pv + pv_tv
    net_d   = debt - (safe(data.get("balance_sheet", {}).get("cash_and_equivalents"), 1) or 0)
    eq_val  = ev - net_d
    shares  = inputs.get("shares_outstanding") or safe(inc.get("shares_diluted"), 2) or 1
    price   = eq_val / shares if shares else None

    return {
        "company": data.get("company_name"),
        "wacc_pct": round(wacc * 100, 2),
        "cost_equity_pct": round(cost_e * 100, 2),
        "terminal_growth": round(tgr * 100, 2),
        "years": years, "revenues": revenues, "ebitdas": ebitdas,
        "fcfs": fcfs, "pv_fcfs": pv_fcfs,
        "terminal_value": round(tv_fcf, 1), "pv_terminal": round(pv_tv, 1),
        "sum_pv_fcf": round(sum_pv, 1), "enterprise_value": round(ev, 1),
        "net_debt": round(net_d, 1), "equity_value": round(eq_val, 1),
        "shares": shares,
        "implied_price": round(price, 2) if price else None,
        "base_revenue": base_rev,
        "ebitda_margin_pct": round(ebitda_m * 100, 1),
    }


# ── Excel builder ──────────────────────────────────────────────────────────────

def _hdr(ws, row, col, value, bold=True, bg=DARK_BLUE, fg=WHITE, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def _val(ws, row, col, value, fmt=None, bold=False, bg=None, fg="000000", indent=0):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="right" if col > 1 else "left", indent=indent)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

def _section_hdr(ws, row, c1, c2, label):
    c = ws.cell(row=row, column=c1, value=label)
    c.font  = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.fill  = PatternFill("solid", fgColor=MID_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)


def build_excel_bytes(extracted, analyses, dcf_results) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    for col in ["B","C","D"]:
        ws.column_dimensions[col].width = 22
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Financial Analysis Summary"
    t.font  = Font(bold=True, size=16, color=WHITE, name="Calibri")
    t.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = f"Generated {datetime.now().strftime('%B %d, %Y')}"
    sub.font  = Font(size=10, color="888888", name="Calibri")
    sub.alignment = Alignment(horizontal="center")

    row = 4
    names = [a["company"] or f"Company {i+1}" for i,a in enumerate(analyses)]
    _hdr(ws, row, 1, "Metric", bg=MID_BLUE)
    for c, nm in enumerate(names, 2):
        _hdr(ws, row, c, nm, bg=MID_BLUE)
    row += 1

    dcf_map = {d["company"]: d for d in dcf_results}
    metrics = [
        ("── Income Statement ──────────────────", None, None),
        ("Revenue (USD M)",    "revenue_y0",         "#,##0.0"),
        ("Revenue YoY %",      "revenue_yoy_pct",    '0.0"%"'),
        ("Gross Margin %",     "gross_margin_pct",   '0.0"%"'),
        ("EBITDA Margin %",    "ebitda_margin_pct",  '0.0"%"'),
        ("Net Margin %",       "net_margin_pct",     '0.0"%"'),
        ("EPS (Diluted)",      "eps_diluted",        "$0.00"),
        ("── Balance Sheet ────────────────────", None, None),
        ("Current Ratio",      "current_ratio",      "0.00"),
        ("Debt / Equity",      "debt_to_equity",     "0.00"),
        ("Net Debt (USD M)",   "net_debt",           "#,##0.0"),
        ("ROE %",              "roe_pct",            '0.0"%"'),
        ("── Cash Flow ────────────────────────", None, None),
        ("FCF (USD M)",        "fcf_y0",             "#,##0.0"),
        ("FCF Margin %",       "fcf_margin",         '0.0"%"'),
        ("── DCF Valuation ───────────────────", None, None),
        ("WACC %",             "wacc_pct",           '0.00"%"'),
        ("Enterprise Value",   "enterprise_value",   "#,##0.0"),
        ("Equity Value",       "equity_value",       "#,##0.0"),
        ("Implied Price",      "implied_price",      "$#,##0.00"),
    ]
    for label, key, fmt in metrics:
        if key is None:
            _section_hdr(ws, row, 1, 4, label); row += 1; continue
        _val(ws, row, 1, label)
        for c, a in enumerate(analyses, 2):
            v = a.get(key)
            if v is None:
                v = dcf_map.get(a["company"], {}).get(key)
            _val(ws, row, c, v, fmt=fmt,
                 fg=(NEG_RED if isinstance(v, (int,float)) and v < 0 else "000000"))
        row += 1

    # ── Per-company sheets ─────────────────────────────────────────────────
    for idx, (data, analysis, dcf) in enumerate(zip(extracted, analyses, dcf_results), 1):
        name  = data.get("company_name") or f"Company {idx}"
        short = re.sub(r"[\\/*?\[\]:]", "", name)[:28]
        ws    = wb.create_sheet(f"{idx}. {short}")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 36
        for col in ["B","C","D","E"]:
            ws.column_dimensions[col].width = 16

        yr  = int(data.get("fiscal_year", datetime.now().year))
        yrs = [yr-2, yr-1, yr]
        inc = data.get("income_statement", {})
        bs  = data.get("balance_sheet", {})
        cf  = data.get("cash_flow", {})

        ws.merge_cells("A1:E1")
        t = ws["A1"]
        t.value = f"{name}  |  FY{yr} Financial Statements"
        t.font  = Font(bold=True, size=14, color=WHITE, name="Calibri")
        t.fill  = PatternFill("solid", fgColor=DARK_BLUE)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32
        row = 3

        def g3(d, key): return [safe(d.get(key), i) for i in range(3)]
        def g2(d, key): return [safe(d.get(key), i) for i in range(2)]
        M = "#,##0.0"; P = '0.0"%"'

        def wr(label, vals, fmt, bold=False, bg=None, indent=0):
            nonlocal row
            _val(ws, row, 1, label, bold=bold, bg=bg, indent=indent)
            for c, v in enumerate(vals, 2):
                _val(ws, row, c, v, fmt=fmt, bold=bold, bg=bg,
                     fg=(NEG_RED if isinstance(v,(int,float)) and v < 0 else "000000"))
            row += 1

        # Income Statement
        _section_hdr(ws, row, 1, 5, "INCOME STATEMENT"); row += 1
        _hdr(ws, row, 1, "USD Millions", bg=MID_BLUE, size=9)
        for c, y in enumerate(yrs, 2): _hdr(ws, row, c, f"FY{y}", bg=MID_BLUE)
        row += 1
        rev = g3(inc, "revenue")
        wr("Revenue",            rev,                      M, bold=True, bg=LIGHT_GREY)
        wr("Cost of Goods Sold", g3(inc,"cost_of_goods_sold"), M, indent=1)
        wr("Gross Profit",       g3(inc,"gross_profit"),   M, bold=True)
        wr("  Gross Margin %",   [pct(g3(inc,"gross_profit")[i], rev[i]) for i in range(3)], P, indent=2)
        wr("R&D Expense",        g3(inc,"rd_expense"),     M, indent=1)
        wr("SG&A Expense",       g3(inc,"sg_a_expense"),   M, indent=1)
        wr("Operating Income",   g3(inc,"operating_income"),M, bold=True)
        wr("EBITDA",             g3(inc,"ebitda"),         M, bold=True, bg=LIGHT_GREY)
        wr("  EBITDA Margin %",  [pct(g3(inc,"ebitda")[i], rev[i]) for i in range(3)], P, indent=2)
        wr("Interest Expense",   g3(inc,"interest_expense"),M, indent=1)
        wr("Net Income",         g3(inc,"net_income"),     M, bold=True, bg=LIGHT_BLUE)
        wr("  Net Margin %",     [pct(g3(inc,"net_income")[i], rev[i]) for i in range(3)], P, indent=2)
        wr("EPS (Diluted)",      g3(inc,"eps_diluted"),    "$0.00")
        row += 1

        # Balance Sheet
        _section_hdr(ws, row, 1, 5, "BALANCE SHEET"); row += 1
        bs_yrs = [yr-1, yr]
        _hdr(ws, row, 1, "USD Millions", bg=MID_BLUE, size=9)
        for c, y in enumerate(bs_yrs, 2): _hdr(ws, row, c, f"FY{y}", bg=MID_BLUE)
        row += 1
        wr("Cash & Equivalents",        g2(bs,"cash_and_equivalents"),    M, indent=1)
        wr("Accounts Receivable",        g2(bs,"accounts_receivable"),     M, indent=1)
        wr("Inventory",                  g2(bs,"inventory"),               M, indent=1)
        wr("Total Current Assets",       g2(bs,"total_current_assets"),    M, bold=True)
        wr("PP&E (net)",                 g2(bs,"ppe_net"),                 M, indent=1)
        wr("Goodwill",                   g2(bs,"goodwill"),                M, indent=1)
        wr("TOTAL ASSETS",              g2(bs,"total_assets"),            M, bold=True, bg=LIGHT_GREY)
        wr("Accounts Payable",           g2(bs,"accounts_payable"),        M, indent=1)
        wr("Long-term Debt",             g2(bs,"long_term_debt"),          M, indent=1)
        wr("TOTAL LIABILITIES",         g2(bs,"total_liabilities"),       M, bold=True, bg=LIGHT_GREY)
        wr("TOTAL EQUITY",              g2(bs,"total_equity"),            M, bold=True, bg=LIGHT_BLUE)
        row += 1

        # Cash Flow
        _section_hdr(ws, row, 1, 5, "CASH FLOW STATEMENT"); row += 1
        _hdr(ws, row, 1, "USD Millions", bg=MID_BLUE, size=9)
        for c, y in enumerate(yrs, 2): _hdr(ws, row, c, f"FY{y}", bg=MID_BLUE)
        row += 1
        wr("Operating Cash Flow",   g3(cf,"cfo"),               M, bold=True)
        wr("  D&A",                 g3(cf,"depreciation_amortization"), M, indent=2)
        wr("Capital Expenditure",   g3(cf,"capex"),             M, indent=1)
        fcf = g3(cf,"free_cash_flow")
        wr("Free Cash Flow",        fcf,                        M, bold=True, bg=LIGHT_BLUE)
        wr("  FCF Margin %",        [pct(fcf[i], rev[i]) for i in range(3)], P, indent=2)
        wr("Financing Cash Flow",   g3(cf,"cff"),               M, bold=True)
        wr("Net Change in Cash",    g3(cf,"net_change_in_cash"),M, bold=True, bg=LIGHT_GREY)
        row += 1

        # DCF
        _section_hdr(ws, row, 1, 5, "DCF VALUATION"); row += 1
        _hdr(ws, row, 1, "USD Millions", bg=MID_BLUE, size=9)
        for c, y in enumerate([f"Y+{i}" for i in dcf["years"]], 2):
            _hdr(ws, row, c, y, bg=MID_BLUE)
        row += 1
        _val(ws, row, 1, f"WACC: {dcf['wacc_pct']}%  |  Terminal Growth: {dcf['terminal_growth']}%  |  EBITDA Margin: {dcf['ebitda_margin_pct']}%",
             bold=True, fg="2f5496"); row += 1
        wr("Projected Revenue", dcf["revenues"][:4], M, bold=True, bg=LIGHT_GREY)
        wr("EBITDA",            dcf["ebitdas"][:4],  M)
        wr("Free Cash Flow",    dcf["fcfs"][:4],     M, bold=True)
        wr("PV of FCF",         dcf["pv_fcfs"][:4],  M)
        row += 1
        for label, val, fmt in [
            ("Sum of PV (FCFs)",    dcf["sum_pv_fcf"],      M),
            ("PV of Terminal Value",dcf["pv_terminal"],      M),
            ("Enterprise Value",    dcf["enterprise_value"], M),
            ("Less: Net Debt",      dcf["net_debt"],         M),
            ("Equity Value",        dcf["equity_value"],     M),
            ("Implied Share Price", dcf["implied_price"],    "$#,##0.00"),
        ]:
            bold = label in ("Enterprise Value","Equity Value","Implied Share Price")
            bg   = LIGHT_BLUE if label == "Implied Share Price" else (LIGHT_GREY if bold else None)
            _val(ws, row, 1, label, bold=bold, bg=bg)
            _val(ws, row, 2, val,   fmt=fmt, bold=bold, bg=bg,
                 fg=(NEG_RED if isinstance(val,(int,float)) and val < 0 else "000000"))
            row += 1

    # ── DCF Comparison ─────────────────────────────────────────────────────
    ws = wb.create_sheet("📈 DCF Comparison")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    for col in ["B","C","D"]: ws.column_dimensions[col].width = 20
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "DCF Valuation Comparison"
    t.font  = Font(bold=True, size=14, color=WHITE, name="Calibri")
    t.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    row = 3
    _hdr(ws, row, 1, "Metric", bg=MID_BLUE)
    for c, d in enumerate(dcf_results, 2):
        _hdr(ws, row, c, d["company"] or f"Co {c-1}", bg=MID_BLUE)
    row += 1
    for label, key, fmt in [
        ("Base Revenue (USD M)",      "base_revenue",      "#,##0.0"),
        ("EBITDA Margin %",           "ebitda_margin_pct", '0.0"%"'),
        ("WACC %",                    "wacc_pct",          '0.00"%"'),
        ("Enterprise Value (USD M)",  "enterprise_value",  "#,##0.0"),
        ("Equity Value (USD M)",      "equity_value",      "#,##0.0"),
        ("Implied Share Price (USD)", "implied_price",     "$#,##0.00"),
    ]:
        bold = key in ("enterprise_value","equity_value","implied_price")
        bg   = LIGHT_BLUE if key == "implied_price" else (LIGHT_GREY if bold else None)
        _val(ws, row, 1, label, bold=bold, bg=bg)
        for c, d in enumerate(dcf_results, 2):
            v = d.get(key)
            _val(ws, row, c, v, fmt=fmt, bold=bold, bg=bg,
                 fg=(NEG_RED if isinstance(v,(int,float)) and v < 0 else "000000"))
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 📊 10-K Financial Analyst")
    st.markdown("---")

    # API Key
    st.markdown("### 🔑 API Key")
    api_key_input = st.text_input(
        "Anthropic API Key",
        value=st.session_state.api_key or os.getenv("ANTHROPIC_API_KEY", ""),
        type="password",
        placeholder="sk-ant-...",
        label_visibility="collapsed",
    )
    if api_key_input:
        st.session_state.api_key = api_key_input

    st.markdown("---")

    # File upload
    st.markdown("### 📁 Upload 10-K Files")
    st.caption("Upload up to 3 company 10-K filings (PDF or TXT)")

    uploaded_files = st.file_uploader(
        "Drop files here",
        type=["pdf", "txt"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    if uploaded_files:
        new_files = [f for f in uploaded_files
                     if f.name not in [c.get("filename") for c in st.session_state.companies]]
        for uf in new_files[:max(0, 3 - len(st.session_state.companies))]:
            with st.spinner(f"Reading {uf.name}…"):
                text = read_uploaded_file(uf)
            # guess company name from filename
            guessed = Path(uf.name).stem.replace("_", " ").replace("-", " ").title()
            st.session_state.companies.append({
                "name": guessed,
                "text": text,
                "filename": uf.name,
                "chars": len(text),
            })

    # Show loaded companies
    st.markdown("---")
    st.markdown("### 🏢 Loaded Companies")
    if not st.session_state.companies:
        st.caption("No companies loaded yet")
    else:
        for i, c in enumerate(st.session_state.companies):
            col1, col2 = st.columns([4, 1])
            with col1:
                new_name = st.text_input(
                    f"Co {i+1}", value=c["name"], key=f"name_{i}",
                    label_visibility="collapsed"
                )
                st.session_state.companies[i]["name"] = new_name
                st.caption(f"{c['chars']:,} chars · {c['filename']}")
            with col2:
                if st.button("✕", key=f"del_{i}", help="Remove"):
                    st.session_state.companies.pop(i)
                    st.rerun()

    st.markdown("---")

    # Run Analysis button
    run_disabled = (
        len(st.session_state.companies) == 0
        or not st.session_state.api_key
    )
    if st.button("🚀 Run Analysis", disabled=run_disabled, use_container_width=True):
        st.session_state.analysis_done = False
        st.session_state.extracted     = []
        st.session_state.analyses      = []
        st.session_state.dcf_results   = []
        st.session_state.excel_bytes   = None
        client = get_client()

        progress = st.progress(0, text="Starting…")
        n = len(st.session_state.companies)

        all_ok = True
        for i, c in enumerate(st.session_state.companies):
            progress.progress((i) / n, text=f"Extracting {c['name']}…")
            try:
                data = extract_financial_data(c["name"], c["text"], client)
                data["company_name"] = c["name"]  # use user-edited name
                st.session_state.extracted.append(data)
                st.session_state.analyses.append(three_statement_analysis(data))
                st.session_state.dcf_results.append(build_dcf(data))
            except Exception as e:
                st.error(f"Error extracting {c['name']}: {e}")
                all_ok = False

        if all_ok and st.session_state.extracted:
            progress.progress(1.0, text="Building Excel…")
            try:
                st.session_state.excel_bytes = build_excel_bytes(
                    st.session_state.extracted,
                    st.session_state.analyses,
                    st.session_state.dcf_results,
                )
            except Exception as e:
                st.warning(f"Excel build failed: {e}")

            st.session_state.analysis_done = True
            # inject summary into chat
            names = [a["company"] for a in st.session_state.analyses]
            st.session_state.messages.append({
                "role": "assistant",
                "content": (
                    f"✅ Analysis complete for **{', '.join(names)}**! "
                    "Scroll down to see the results, and download the Excel report above. "
                    "Ask me anything about the financials!"
                ),
            })
            progress.empty()
            st.rerun()

    # Excel download
    if st.session_state.excel_bytes:
        st.download_button(
            label="⬇️ Download Excel Report",
            data=st.session_state.excel_bytes,
            file_name=f"financial_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # Clear
    if st.button("🗑️ Clear All", use_container_width=True):
        for k in ["companies", "extracted", "analyses", "dcf_results",
                  "excel_bytes", "messages"]:
            st.session_state[k] = [] if k != "excel_bytes" else None
        st.session_state.analysis_done = False
        st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
# MAIN AREA
# ═════════════════════════════════════════════════════════════════════════════

st.markdown("## 📊 10-K Financial Analysis Chatbot")
st.caption("Upload 10-K filings → Run Analysis → Chat about the results")

# ── Results dashboard ─────────────────────────────────────────────────────────
if st.session_state.analysis_done and st.session_state.analyses:
    cols = st.columns(len(st.session_state.analyses))
    for col, a, dcf in zip(cols, st.session_state.analyses, st.session_state.dcf_results):
        with col:
            nm = a.get("company") or "Company"
            st.markdown(f"""
            <div class="analysis-card">
              <b style="font-size:15px;color:#1f3864">{nm}</b>
              <div class="metric-row">
                <div class="metric">
                  <div class="metric-val">${a.get('revenue_y0') or '—':.0f}M</div>
                  <div class="metric-lbl">Revenue</div>
                </div>
                <div class="metric">
                  <div class="metric-val">{a.get('ebitda_margin_pct') or '—'}%</div>
                  <div class="metric-lbl">EBITDA Margin</div>
                </div>
                <div class="metric">
                  <div class="metric-val">${dcf.get('implied_price') or '—'}</div>
                  <div class="metric-lbl">Implied Price</div>
                </div>
                <div class="metric">
                  <div class="metric-val">{dcf.get('wacc_pct') or '—'}%</div>
                  <div class="metric-lbl">WACC</div>
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

    # DCF table
    with st.expander("📈 DCF Comparison Table", expanded=False):
        dcf_table = {}
        for dcf in st.session_state.dcf_results:
            nm = dcf.get("company") or "?"
            dcf_table[nm] = {
                "WACC %":             f"{dcf.get('wacc_pct')}%",
                "Enterprise Value":   f"${dcf.get('enterprise_value'):,.0f}M",
                "Equity Value":       f"${dcf.get('equity_value'):,.0f}M",
                "Implied Price":      f"${dcf.get('implied_price')}",
                "Terminal Growth %":  f"{dcf.get('terminal_growth')}%",
            }
        import pandas as pd
        st.dataframe(pd.DataFrame(dcf_table).T, use_container_width=True)

    st.markdown("---")

# ── Welcome message ───────────────────────────────────────────────────────────
if not st.session_state.messages:
    with st.chat_message("assistant", avatar="📊"):
        st.markdown("""
        Welcome! I'm your **10-K Financial Analysis Assistant**.

        **How to get started:**
        1. 🔑 Enter your Anthropic API key in the sidebar
        2. 📁 Upload 1–3 company 10-K filings (PDF or TXT) in the sidebar
        3. 🚀 Click **Run Analysis**
        4. 💬 Ask me anything about the results!

        **Example questions after analysis:**
        - *"Which company has the best FCF margin?"*
        - *"Compare the WACC and valuation assumptions"*
        - *"Which stock looks most undervalued based on the DCF?"*
        - *"What are the key risks for each company?"*
        """)

# ── Chat history ──────────────────────────────────────────────────────────────
for msg in st.session_state.messages:
    avatar = "📊" if msg["role"] == "assistant" else "🧑"
    with st.chat_message(msg["role"], avatar=avatar):
        st.markdown(msg["content"])

# ── Chat input ────────────────────────────────────────────────────────────────
if prompt := st.chat_input("Ask about the financials…"):
    if not st.session_state.api_key:
        st.warning("Please enter your Anthropic API key in the sidebar.")
        st.stop()

    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user", avatar="🧑"):
        st.markdown(prompt)

    # Build context
    system_ctx = "You are a senior equity research analyst."
    if st.session_state.analysis_done and st.session_state.extracted:
        ctx_json = json.dumps(st.session_state.extracted, indent=2)[:40_000]
        system_ctx = (
            "You are a senior equity research analyst. "
            "Here is the extracted financial data:\n" + ctx_json
        )

    with st.chat_message("assistant", avatar="📊"):
        placeholder = st.empty()
        client = get_client()
        full_reply = ""
        with client.messages.stream(
            model=MODEL,
            max_tokens=2048,
            system=system_ctx,
            messages=[
                {"role": m["role"], "content": m["content"]}
                for m in st.session_state.messages
            ],
        ) as stream:
            for text in stream.text_stream:
                full_reply += text
                placeholder.markdown(full_reply + "▌")
        placeholder.markdown(full_reply)

    st.session_state.messages.append({"role": "assistant", "content": full_reply})
